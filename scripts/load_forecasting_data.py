#!/usr/bin/env python3
"""Load yearly roaming Excel files into the SQLite forecasting database.

Pipeline step 1: reads *Actual*.xlsx workbooks, auto-detects the header row
matching EXPECTED_COLUMNS, converts values to proper types, and bulk-inserts
into the 'traffic' table in data/forecasting.db. Also creates a
'traffic_features' view with derived year/month columns, and indexes for
fast querying.

The schema has 28 columns: 9 text (identifiers) + 18 numeric (traffic metrics)
+ 1 CALL_YEAR_MONTH (integer YYYYMM). Numeric columns cover inbound/outbound
calls, volume, duration, charges, and taxes.

Usage:
  python scripts/load_forecasting_data.py --replace   # fresh load
  python scripts/load_forecasting_data.py             # append to existing
"""
from __future__ import annotations

import argparse
import sqlite3
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


TABLE_NAME = "traffic"
VIEW_NAME = "traffic_features"

EXPECTED_COLUMNS = [
    "TRAFFIC_TYPE",
    "CALL_YEAR_MONTH",
    "SRC_TADIG",
    "CALL_TYPE",
    "DST_TADIG",
    "DST_NAME",
    "DST_COUNTRY",
    "GROUPNAME",
    "NEGOTIATOR",
    "INBOUND_CALLS",
    "INBOUND_VOL_MB",
    "INBOUND_DURATION",
    "INBOUND_CHARGED_VOLUME_MB",
    "INBOUND_CHARGED_DURATION",
    "INBOUND_CHARGES_EUR",
    "INBOUND_TAXES_EUR",
    "INBOUND_CHARGES_SDR",
    "INBOUND_TAXES_SDR",
    "OUTBOUND_CALLS",
    "OUTBOUND_VOL_MB",
    "OUTBOUND_DURATION",
    "OUTBOUND_CHARGED_VOLUME_MB",
    "OUTBOUND_CHARGED_DURATION",
    "OUTBOUND_CHARGES_EUR",
    "OUTBOUND_TAXES_EUR",
    "OUTBOUND_CHARGES_SDR",
    "OUTBOUND_TAXES_SDR",
    "CALL_DESTINATION",
]

TEXT_COLUMNS = {
    "TRAFFIC_TYPE",
    "SRC_TADIG",
    "CALL_TYPE",
    "DST_TADIG",
    "DST_NAME",
    "DST_COUNTRY",
    "GROUPNAME",
    "NEGOTIATOR",
    "CALL_DESTINATION",
}

NUMERIC_COLUMNS = {
    "INBOUND_CALLS",
    "INBOUND_VOL_MB",
    "INBOUND_DURATION",
    "INBOUND_CHARGED_VOLUME_MB",
    "INBOUND_CHARGED_DURATION",
    "INBOUND_CHARGES_EUR",
    "INBOUND_TAXES_EUR",
    "INBOUND_CHARGES_SDR",
    "INBOUND_TAXES_SDR",
    "OUTBOUND_CALLS",
    "OUTBOUND_VOL_MB",
    "OUTBOUND_DURATION",
    "OUTBOUND_CHARGED_VOLUME_MB",
    "OUTBOUND_CHARGED_DURATION",
    "OUTBOUND_CHARGES_EUR",
    "OUTBOUND_TAXES_EUR",
    "OUTBOUND_CHARGES_SDR",
    "OUTBOUND_TAXES_SDR",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Load yearly Excel files into SQLite for forecasting."
    )
    parser.add_argument(
        "--pattern",
        default="*Actual*.xlsx",
        help="Glob pattern to locate Excel files.",
    )
    parser.add_argument(
        "--db-path",
        default="data/forecasting.db",
        help="Output SQLite path.",
    )
    parser.add_argument(
        "--sheet-name",
        default=None,
        help="Optional fixed sheet name. By default, script auto-detects sheet by header.",
    )
    parser.add_argument(
        "--header-scan-rows",
        type=int,
        default=40,
        help="Rows to scan when searching for header.",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=5000,
        help="Batch size for inserts.",
    )
    parser.add_argument(
        "--replace",
        action="store_true",
        help="Drop and recreate table before load.",
    )
    parser.add_argument(
        "--empty-row-stop",
        type=int,
        default=1000,
        help="Stop sheet scan after this many consecutive empty rows.",
    )
    parser.add_argument(
        "--report-every",
        type=int,
        default=50000,
        help="Print progress every N inserted rows per file.",
    )
    return parser.parse_args()


def normalize_header_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().upper()


def find_header_row(ws, header_scan_rows: int) -> tuple[int, list[str]]:
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=header_scan_rows, values_only=True), start=1
    ):
        normalized = [normalize_header_cell(v) for v in row]
        while normalized and normalized[-1] == "":
            normalized.pop()
        if normalized == EXPECTED_COLUMNS:
            return row_idx, normalized
    raise ValueError("Could not find expected header row in scanned range.")


def choose_worksheet(wb, fixed_name: str | None, header_scan_rows: int):
    if fixed_name:
        if fixed_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{fixed_name}' not found. Available: {wb.sheetnames}")
        ws = wb[fixed_name]
        header_row, _ = find_header_row(ws, header_scan_rows)
        return ws, header_row

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            header_row, _ = find_header_row(ws, header_scan_rows)
            return ws, header_row
        except ValueError:
            continue
    raise ValueError("No sheet with expected columns found.")


def to_nullable_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text != "" else None


def to_nullable_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text == "":
        return None
    if "," in text and "." in text:
        text = text.replace(",", "")
    elif "," in text and "." not in text:
        text = text.replace(",", ".")
    return float(text)


def to_nullable_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip()
    if text == "":
        return None
    return int(float(text))


def convert_value(column: str, value: Any) -> Any:
    if column == "CALL_YEAR_MONTH":
        return to_nullable_int(value)
    if column in NUMERIC_COLUMNS:
        return to_nullable_float(value)
    if column in TEXT_COLUMNS:
        return to_nullable_text(value)
    return value


def create_schema(conn: sqlite3.Connection, replace: bool) -> None:
    cur = conn.cursor()

    if replace:
        cur.execute(f"DROP VIEW IF EXISTS {VIEW_NAME}")
        drop_indexes(cur)
        cur.execute(f"DROP TABLE IF EXISTS {TABLE_NAME}")

    cur.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
            TRAFFIC_TYPE TEXT,
            CALL_YEAR_MONTH INTEGER NOT NULL,
            SRC_TADIG TEXT,
            CALL_TYPE TEXT,
            DST_TADIG TEXT,
            DST_NAME TEXT,
            DST_COUNTRY TEXT,
            GROUPNAME TEXT,
            NEGOTIATOR TEXT,
            INBOUND_CALLS REAL,
            INBOUND_VOL_MB REAL,
            INBOUND_DURATION REAL,
            INBOUND_CHARGED_VOLUME_MB REAL,
            INBOUND_CHARGED_DURATION REAL,
            INBOUND_CHARGES_EUR REAL,
            INBOUND_TAXES_EUR REAL,
            INBOUND_CHARGES_SDR REAL,
            INBOUND_TAXES_SDR REAL,
            OUTBOUND_CALLS REAL,
            OUTBOUND_VOL_MB REAL,
            OUTBOUND_DURATION REAL,
            OUTBOUND_CHARGED_VOLUME_MB REAL,
            OUTBOUND_CHARGED_DURATION REAL,
            OUTBOUND_CHARGES_EUR REAL,
            OUTBOUND_TAXES_EUR REAL,
            OUTBOUND_CHARGES_SDR REAL,
            OUTBOUND_TAXES_SDR REAL,
            CALL_DESTINATION TEXT
        )
        """
    )

    cur.execute(
        f"""
        CREATE VIEW IF NOT EXISTS {VIEW_NAME} AS
        SELECT
            *,
            CAST(SUBSTR(CAST(CALL_YEAR_MONTH AS TEXT), 1, 4) AS INTEGER) AS CALL_YEAR,
            CAST(SUBSTR(CAST(CALL_YEAR_MONTH AS TEXT), 5, 2) AS INTEGER) AS CALL_MONTH,
            DATE(
                SUBSTR(CAST(CALL_YEAR_MONTH AS TEXT), 1, 4) || '-' ||
                SUBSTR(CAST(CALL_YEAR_MONTH AS TEXT), 5, 2) || '-01'
            ) AS MONTH_START
        FROM {TABLE_NAME}
        """
    )
    conn.commit()


def drop_indexes(cur: sqlite3.Cursor) -> None:
    cur.execute(f"DROP INDEX IF EXISTS idx_{TABLE_NAME}_month")
    cur.execute(f"DROP INDEX IF EXISTS idx_{TABLE_NAME}_series")
    cur.execute(f"DROP INDEX IF EXISTS idx_{TABLE_NAME}_group_month")


def create_indexes(conn: sqlite3.Connection) -> None:
    cur = conn.cursor()
    cur.execute(f"CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_month ON {TABLE_NAME}(CALL_YEAR_MONTH)")
    cur.execute(
        f"""
        CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_series
        ON {TABLE_NAME}(SRC_TADIG, DST_TADIG, CALL_TYPE, CALL_DESTINATION, CALL_YEAR_MONTH)
        """
    )
    cur.execute(
        f"""
        CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_group_month
        ON {TABLE_NAME}(GROUPNAME, NEGOTIATOR, CALL_YEAR_MONTH)
        """
    )
    conn.commit()


def configure_for_fast_load(conn: sqlite3.Connection) -> None:
    cur = conn.cursor()
    cur.execute("PRAGMA journal_mode=MEMORY;")
    cur.execute("PRAGMA synchronous=OFF;")
    cur.execute("PRAGMA temp_store=MEMORY;")
    cur.execute("PRAGMA locking_mode=EXCLUSIVE;")
    cur.execute("PRAGMA cache_size=-200000;")


def configure_for_normal_ops(conn: sqlite3.Connection) -> None:
    cur = conn.cursor()
    cur.execute("PRAGMA locking_mode=NORMAL;")
    cur.execute("PRAGMA journal_mode=WAL;")
    cur.execute("PRAGMA synchronous=NORMAL;")


def load_excel_file(
    conn: sqlite3.Connection,
    excel_path: Path,
    sheet_name: str | None,
    header_scan_rows: int,
    batch_size: int,
    empty_row_stop: int,
    report_every: int,
) -> int:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        ws, header_row = choose_worksheet(wb, sheet_name, header_scan_rows)
        insert_sql = f"""
            INSERT INTO {TABLE_NAME} (
                {", ".join(EXPECTED_COLUMNS)}
            ) VALUES (
                {", ".join(["?"] * len(EXPECTED_COLUMNS))}
            )
        """
        batch: list[tuple[Any, ...]] = []
        loaded = 0
        empty_streak = 0

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            values = list(row[: len(EXPECTED_COLUMNS)])
            if len(values) < len(EXPECTED_COLUMNS):
                values.extend([None] * (len(EXPECTED_COLUMNS) - len(values)))
            if not any(v is not None and str(v).strip() != "" for v in values):
                empty_streak += 1
                if empty_row_stop > 0 and empty_streak >= empty_row_stop:
                    break
                continue

            empty_streak = 0
            converted = tuple(convert_value(col, values[idx]) for idx, col in enumerate(EXPECTED_COLUMNS))
            if converted[1] is None:
                continue
            batch.append(converted)
            loaded += 1

            if len(batch) >= batch_size:
                conn.executemany(insert_sql, batch)
                batch.clear()
                if report_every > 0 and loaded % report_every == 0:
                    print(f"{excel_path.name}: {loaded} rows inserted...")

        if batch:
            conn.executemany(insert_sql, batch)
        return loaded
    finally:
        wb.close()


def main() -> None:
    args = parse_args()
    files = sorted(Path(".").glob(args.pattern))
    if not files:
        raise SystemExit(f"No Excel files found for pattern: {args.pattern}")

    db_path = Path(args.db_path)
    db_path.parent.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(db_path)
    try:
        create_schema(conn, replace=args.replace)
        configure_for_fast_load(conn)
        # Avoid index maintenance during insert-heavy phase.
        drop_indexes(conn.cursor())
        conn.commit()

        conn.execute("BEGIN")
        total_rows = 0
        for file_path in files:
            loaded = load_excel_file(
                conn=conn,
                excel_path=file_path,
                sheet_name=args.sheet_name,
                header_scan_rows=args.header_scan_rows,
                batch_size=args.batch_size,
                empty_row_stop=args.empty_row_stop,
                report_every=args.report_every,
            )
            total_rows += loaded
            print(f"{file_path.name}: loaded {loaded} rows")

        conn.commit()
        configure_for_normal_ops(conn)
        create_indexes(conn)

        count = conn.execute(f"SELECT COUNT(*) FROM {TABLE_NAME}").fetchone()[0]
        print(f"Total loaded this run: {total_rows}")
        print(f"Total rows in table {TABLE_NAME}: {count}")
        print(f"SQLite database ready: {db_path}")
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
