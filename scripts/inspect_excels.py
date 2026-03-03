#!/usr/bin/env python3
"""Inspect and optionally load raw Excel workbooks into SQLite.

Utility for initial data ingestion and schema discovery. Scans Excel files
matching a glob pattern, profiles each sheet (header detection, column listing,
data row counts), and optionally loads cell-level data into a SQLite database.

Two modes:
  1. Inspect only (default): produces reports/excel_inventory.json with
     sheet profiles and cross-year schema comparison.
  2. --load-sqlite: additionally loads normalized cell-level records into
     data/forecasting.db (tables: sheet_profiles, cell_values).

This is a one-time setup tool — re-run only when source Excel files change.

Usage:
  python scripts/inspect_excels.py                           # inspect only
  python scripts/inspect_excels.py --load-sqlite --replace   # inspect + load
"""
from __future__ import annotations

import argparse
import json
import re
import sqlite3
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass
class SheetProfile:
    workbook: str
    year: int | None
    sheet: str
    header_row: int | None
    columns: list[str]
    preview_rows: list[dict[str, Any]]
    data_rows_scanned: int
    max_row: int | None
    max_column: int | None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Inspect Excel files, compare sheet columns across years, and optionally "
            "load normalized records into SQLite."
        )
    )
    parser.add_argument("--pattern", default="*Actual*.xlsx", help="Glob pattern for workbooks.")
    parser.add_argument(
        "--preview-rows", type=int, default=5, help="Number of data rows to sample per sheet."
    )
    parser.add_argument(
        "--header-scan-rows",
        type=int,
        default=40,
        help="Rows to scan before deciding where the header likely starts.",
    )
    parser.add_argument(
        "--output-json",
        default="reports/excel_inventory.json",
        help="Path for machine-readable inventory output.",
    )
    parser.add_argument(
        "--load-sqlite",
        action="store_true",
        help="Load normalized cell-level data into SQLite.",
    )
    parser.add_argument(
        "--sqlite-path",
        default="data/forecasting.db",
        help="SQLite DB path (used when --load-sqlite is set).",
    )
    parser.add_argument(
        "--max-data-rows-per-sheet",
        type=int,
        default=None,
        help="Optional cap for data rows loaded per sheet (after header).",
    )
    parser.add_argument(
        "--empty-row-stop",
        type=int,
        default=2000,
        help="Stop scanning a sheet after this many consecutive empty rows after header.",
    )
    return parser.parse_args()


def parse_year_from_name(name: str) -> int | None:
    match = re.search(r"(19|20)\d{2}", name)
    return int(match.group(0)) if match else None


def normalize_header_value(value: Any, idx: int) -> str:
    if value is None:
        return f"col_{idx}"
    text = str(value).strip()
    return text if text else f"col_{idx}"


def dedupe_columns(columns: list[str]) -> list[str]:
    counts: dict[str, int] = defaultdict(int)
    output: list[str] = []
    for col in columns:
        counts[col] += 1
        if counts[col] == 1:
            output.append(col)
        else:
            output.append(f"{col}__{counts[col]}")
    return output


def trim_trailing_empty(row: tuple[Any, ...]) -> list[Any]:
    values = list(row)
    while values and (values[-1] is None or values[-1] == ""):
        values.pop()
    return values


def is_likely_header(values: list[Any]) -> bool:
    non_empty = [v for v in values if v is not None and str(v).strip() != ""]
    if len(non_empty) < 2:
        return False
    text_like = sum(isinstance(v, str) for v in non_empty)
    return text_like >= 1


def excel_value_to_sql(v: Any) -> tuple[str, str | None, float | None]:
    if v is None:
        return ("null", None, None)
    if isinstance(v, bool):
        return ("bool", "1" if v else "0", 1.0 if v else 0.0)
    if isinstance(v, (int, float)):
        return ("number", str(v), float(v))
    if isinstance(v, datetime):
        return ("datetime", v.isoformat(sep=" "), None)
    if isinstance(v, date):
        return ("date", v.isoformat(), None)
    if isinstance(v, time):
        return ("time", v.isoformat(), None)
    return ("text", str(v), None)


def inspect_sheet(
    workbook_name: str,
    year: int | None,
    sheet_name: str,
    ws,
    preview_rows: int,
    header_scan_rows: int,
    sqlite_conn: sqlite3.Connection | None = None,
    max_data_rows_per_sheet: int | None = None,
    empty_row_stop: int = 2000,
) -> SheetProfile:
    header_row_idx: int | None = None
    columns: list[str] = []
    preview: list[dict[str, Any]] = []
    data_rows_scanned = 0
    empty_streak = 0
    inserted_rows: list[tuple[Any, ...]] = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        trimmed = trim_trailing_empty(row)

        if header_row_idx is None and row_idx <= header_scan_rows:
            if is_likely_header(trimmed):
                header_row_idx = row_idx
                raw_cols = [normalize_header_value(v, i + 1) for i, v in enumerate(trimmed)]
                columns = dedupe_columns(raw_cols)
            continue

        if header_row_idx is None:
            continue
        if row_idx <= header_row_idx:
            continue

        if max_data_rows_per_sheet is not None and data_rows_scanned >= max_data_rows_per_sheet:
            break

        row_values = trimmed[: len(columns)]
        if not any(v is not None and str(v).strip() != "" for v in row_values):
            empty_streak += 1
            if empty_row_stop > 0 and empty_streak >= empty_row_stop:
                break
            continue

        empty_streak = 0
        data_rows_scanned += 1

        if len(preview) < preview_rows:
            row_map = {}
            for cidx, col_name in enumerate(columns):
                row_map[col_name] = row_values[cidx] if cidx < len(row_values) else None
            preview.append(row_map)

        if sqlite_conn is not None:
            source_row = row_idx
            for cidx, col_name in enumerate(columns):
                val = row_values[cidx] if cidx < len(row_values) else None
                if val is None or str(val).strip() == "":
                    continue
                value_type, value_text, value_num = excel_value_to_sql(val)
                inserted_rows.append(
                    (
                        workbook_name,
                        year,
                        sheet_name,
                        source_row,
                        col_name,
                        value_type,
                        value_text,
                        value_num,
                    )
                )
            if len(inserted_rows) >= 5000:
                sqlite_conn.executemany(
                    """
                    INSERT INTO cell_values
                    (workbook, year, sheet, row_index, column_name, value_type, value_text, value_num)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    inserted_rows,
                )
                inserted_rows.clear()

    if sqlite_conn is not None and inserted_rows:
        sqlite_conn.executemany(
            """
            INSERT INTO cell_values
            (workbook, year, sheet, row_index, column_name, value_type, value_text, value_num)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            inserted_rows,
        )

    return SheetProfile(
        workbook=workbook_name,
        year=year,
        sheet=sheet_name,
        header_row=header_row_idx,
        columns=columns,
        preview_rows=preview,
        data_rows_scanned=data_rows_scanned,
        max_row=ws.max_row,
        max_column=ws.max_column,
    )


def ensure_sqlite_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        PRAGMA journal_mode=WAL;
        CREATE TABLE IF NOT EXISTS sheet_profiles (
            workbook TEXT NOT NULL,
            year INTEGER,
            sheet TEXT NOT NULL,
            header_row INTEGER,
            num_columns INTEGER NOT NULL,
            data_rows_scanned INTEGER NOT NULL,
            max_row INTEGER,
            max_column INTEGER,
            columns_json TEXT NOT NULL,
            preview_json TEXT NOT NULL,
            PRIMARY KEY (workbook, sheet)
        );

        CREATE TABLE IF NOT EXISTS cell_values (
            workbook TEXT NOT NULL,
            year INTEGER,
            sheet TEXT NOT NULL,
            row_index INTEGER NOT NULL,
            column_name TEXT NOT NULL,
            value_type TEXT NOT NULL,
            value_text TEXT,
            value_num REAL
        );

        CREATE INDEX IF NOT EXISTS idx_cell_values_sheet_year
            ON cell_values(sheet, year);
        CREATE INDEX IF NOT EXISTS idx_cell_values_col
            ON cell_values(column_name);
        """
    )
    conn.commit()


def compare_schemas(profiles: list[SheetProfile]) -> dict[str, Any]:
    by_sheet_year: dict[str, dict[int | None, list[str]]] = defaultdict(dict)
    for p in profiles:
        by_sheet_year[p.sheet][p.year] = p.columns

    all_years = sorted({p.year for p in profiles if p.year is not None})
    mismatches: list[dict[str, Any]] = []
    only_in_some_years: list[dict[str, Any]] = []

    for sheet, year_map in sorted(by_sheet_year.items()):
        years_present = sorted(y for y in year_map.keys() if y is not None)
        if years_present != all_years:
            only_in_some_years.append({"sheet": sheet, "years_present": years_present})

        if not years_present:
            continue
        base_year = years_present[0]
        base_cols = year_map[base_year]
        base_set = set(base_cols)

        for year in years_present[1:]:
            current_cols = year_map[year]
            current_set = set(current_cols)
            missing = sorted(base_set - current_set)
            added = sorted(current_set - base_set)
            if missing or added or base_cols != current_cols:
                mismatches.append(
                    {
                        "sheet": sheet,
                        "base_year": base_year,
                        "compare_year": year,
                        "base_col_count": len(base_cols),
                        "compare_col_count": len(current_cols),
                        "missing_vs_base": missing,
                        "added_vs_base": added,
                        "order_changed": (not missing and not added and base_cols != current_cols),
                    }
                )

    return {
        "years_detected": all_years,
        "sheet_count": len(by_sheet_year),
        "mismatch_count": len(mismatches),
        "mismatches": mismatches,
        "only_in_some_years": only_in_some_years,
    }


def main() -> None:
    args = parse_args()
    files = sorted(Path(".").glob(args.pattern))
    if not files:
        raise SystemExit(f"No files found for pattern: {args.pattern}")

    Path(args.output_json).parent.mkdir(parents=True, exist_ok=True)

    sqlite_conn: sqlite3.Connection | None = None
    if args.load_sqlite:
        db_path = Path(args.sqlite_path)
        db_path.parent.mkdir(parents=True, exist_ok=True)
        sqlite_conn = sqlite3.connect(db_path)
        ensure_sqlite_schema(sqlite_conn)
        # Full refresh for deterministic runs
        sqlite_conn.execute("DELETE FROM sheet_profiles")
        sqlite_conn.execute("DELETE FROM cell_values")
        sqlite_conn.commit()

    all_profiles: list[SheetProfile] = []

    for file_path in files:
        year = parse_year_from_name(file_path.name)
        wb = load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            profile = inspect_sheet(
                workbook_name=file_path.name,
                year=year,
                sheet_name=sheet_name,
                ws=ws,
                preview_rows=args.preview_rows,
                header_scan_rows=args.header_scan_rows,
                sqlite_conn=sqlite_conn,
                max_data_rows_per_sheet=args.max_data_rows_per_sheet,
                empty_row_stop=args.empty_row_stop,
            )
            all_profiles.append(profile)

            if sqlite_conn is not None:
                sqlite_conn.execute(
                    """
                    INSERT INTO sheet_profiles
                    (workbook, year, sheet, header_row, num_columns, data_rows_scanned,
                     max_row, max_column, columns_json, preview_json)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        profile.workbook,
                        profile.year,
                        profile.sheet,
                        profile.header_row,
                        len(profile.columns),
                        profile.data_rows_scanned,
                        profile.max_row,
                        profile.max_column,
                        json.dumps(profile.columns, ensure_ascii=True),
                        json.dumps(profile.preview_rows, default=str, ensure_ascii=True),
                    ),
                )
        wb.close()
        if sqlite_conn is not None:
            sqlite_conn.commit()

    comparison = compare_schemas(all_profiles)

    inventory = {
        "files": [f.name for f in files],
        "profiles": [
            {
                "workbook": p.workbook,
                "year": p.year,
                "sheet": p.sheet,
                "header_row": p.header_row,
                "num_columns": len(p.columns),
                "columns": p.columns,
                "preview_rows": p.preview_rows,
                "data_rows_scanned": p.data_rows_scanned,
                "max_row": p.max_row,
                "max_column": p.max_column,
            }
            for p in all_profiles
        ],
        "schema_comparison": comparison,
    }

    with open(args.output_json, "w", encoding="utf-8") as f:
        json.dump(inventory, f, indent=2, default=str, ensure_ascii=True)

    if sqlite_conn is not None:
        sqlite_conn.commit()
        sqlite_conn.close()

    by_workbook: dict[str, int] = defaultdict(int)
    for p in all_profiles:
        by_workbook[p.workbook] += 1

    print("Workbook inspection complete.")
    for wb_name, sheet_count in sorted(by_workbook.items()):
        print(f"- {wb_name}: {sheet_count} sheets")
    print(
        "Schema comparison: "
        f"{comparison['mismatch_count']} mismatches across {comparison['sheet_count']} sheet names."
    )
    if comparison["only_in_some_years"]:
        print(f"Sheets missing in some years: {len(comparison['only_in_some_years'])}")
    print(f"Inventory written to: {args.output_json}")
    if args.load_sqlite:
        print(f"SQLite written to: {args.sqlite_path}")


if __name__ == "__main__":
    main()
